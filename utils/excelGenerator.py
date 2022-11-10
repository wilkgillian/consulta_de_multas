from openpyxl import load_workbook


async def excel_generator_multas(placa: str, multa: str, local_data: str, condutor: str, path: str, orgao: str):
    planilha = load_workbook(path)
    try:
        page_docs = planilha['Multas']
        page_docs.append([placa, multa, local_data, condutor, orgao])
        planilha.save(filename=path)
    except:
        planilha.create_sheet("Multas")
        page_docs = planilha['Multas']
        page_docs.append([placa, multa, local_data, condutor, orgao])
        planilha.save(filename=path)
    pass


async def excel_generator_licenciamento(placa: str, licenciamento: str, data_vencimento: str, path: str):
    planilha = load_workbook(path)
    try:
        page_docs = planilha['Licenciamento']
        page_docs.append([placa, licenciamento, data_vencimento])
        planilha.save(filename=path)
    except:
        planilha.create_sheet("Licenciamento")
        page_docs = planilha['Licenciamento']
        page_docs.append([placa, licenciamento, data_vencimento])
        planilha.save(filename=path)
    pass


async def excel_generator_condutores(motorista: str, data_vencimento: str, situacao: str, path: str):
    planilha = load_workbook(path)
    try:
        page_docs = planilha['Motoristas']
        page_docs.append([motorista, data_vencimento, situacao])
        planilha.save(filename=path)
    except:
        planilha.create_sheet("Motoristas")
        page_docs = planilha['Motoristas']
        page_docs.append([motorista, data_vencimento, situacao])
        planilha.save(filename=path)
    pass
