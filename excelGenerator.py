from openpyxl import load_workbook


async def excel_generator(placa: str, multa: str, local_data: str, condutor: str, path: str, orgao: str):
    planilha = load_workbook(path)
    try:
        page_docs = planilha['Multas']
        page_docs.append([placa, multa, local_data, condutor, orgao])
        planilha.save(filename=path)
    except:
        planilha.create_sheet("Multas")
        page_docs = planilha['Multas']
        page_docs.append([placa, multa, local_data, condutor, orgao])
        planilha.save(filename=path)
    pass
