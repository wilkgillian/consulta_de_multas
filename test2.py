import pyautogui
import time
from datetime import datetime
from playwright.sync_api import sync_playwright
import pandas as pd
from bs4 import BeautifulSoup
import re
import openpyxl

t1 = time.time()
dataAtual = datetime.today()
formatData = dataAtual.strftime("%d/%m/%Y").replace("/", "-")
planilha = pd.read_excel("C:/Users/wilk.silva/Downloads/Controle - FrotaO.xlsx", "Vencimento Documentação", skiprows=1, usecols=['PLACA', 'RENAVAN'])
planilha.to_excel("C:/Users/wilk.silva/Downloads/Consulta dia "+formatData+".xlsx")
planilha_formatada = pd.read_excel("C:/Users/wilk.silva/Downloads/Consulta dia "+formatData+".xlsx")
for index,row in planilha_formatada.iterrows():
    with sync_playwright() as p:
      browser = p.chromium.launch(headless=False, timeout=1000)
      context = browser.new_context()
      page = context.new_page()
      page.goto("https://www.detran.mt.gov.br/")
      try:
        placa = row["PLACA"]
        renavan = row["RENAVAN"]
        try:
          page.locator('.closer').click()
        except:
          continue
        page.locator('#input_placa').fill(str(placa))
        page.locator('#input_renavam').fill(str(renavan).replace(".0",""))
        with context.expect_event("page") as event_info:
          page.locator('//*[@id="formVeiculo"]/div[4]/input[2]').click()
        newTab = event_info.value
        try: 
          time.sleep(1)
          localizador = pyautogui.locateOnScreen('debitos.png', confidence=0.9)
          if (localizador == None):
           select = newTab.locator('//*[@id="cmbTipoDebito"]').click()
           pyautogui.press('down')
           pyautogui.press('down')
           pyautogui.press('enter')
           time.sleep(1)
           infracoes = pyautogui.locateOnScreen('infracoes.png', confidence=0.9)
           if (infracoes == None):
            url = newTab.inner_html('//*[@id="div_servicos_Autuacoes"]/table/tbody/tr[2]/td[2]')
            print("Existem dados")
            soup = BeautifulSoup(url, 'html.parser')
            debitos = soup.find_all('td', attrs={'class': False, 'width': False, 'colspan':False})
            print('--- debitos identificados --- \n')
            print(debitos)
            iContent=0
            placaArray = [placa]
            placaParte1 = placa[0:3]
            placaParte2 =  placa[3:]
            placaReplaced = placaParte1+"-"+placaParte2
            book = openpyxl.load_workbook(filename="C:/Users/wilk.silva/Downloads/Consulta dia "+formatData+".xlsx")
            book.create_sheet("Multas")
            page_multas = book['Multas']
            while iContent < len(debitos):
              results = debitos[iContent].text
              if re.search("às", results):
                  print("\n\nLocal, data, e horário ------->>> "+results+"")
                  findData = re.search("[0-9][0-9]/[0-9][0-9]/[0-9][0-9][0-9][0-9]", results).group(0)
                  findHour = re.search("[0-9][0-9]:[0-9][0-9]", results).group(0)
                  data_hour_for_connecta = str(findData)+" "+str(findHour)
                  convertHour = str(findHour)
                  newHour = convertHour.replace(":", "")
                  aspsHour = newHour.replace(newHour,"'"+newHour+"'")
                  convertHour1 = aspsHour[1:3]
                  secondPartHour = aspsHour[3:5]
                  if (re.compile(convertHour1).match("00")):
                    acres = "0.1"
                    sumHour = int(convertHour1)+float(acres)
                  else:
                    sumHour = int(convertHour1)+1
                  horaAidicionada = str(sumHour)+":"+str(secondPartHour)
                  data_hour_adicionada = str(findData)+ " " +str(horaAidicionada)
                  connecta = context.new_page()
                  connecta.goto("http://www16.itrack.com.br/cmatrix/controlemonitoramento")
                  connecta.locator("input[name='usuario']").fill("GEAD")
                  connecta.locator("input[name='senha']").fill("33312976")
                  connecta.locator("button[name='Submit']").click()
                  connecta.locator("//*[@id='tabelaMenu']/tbody/tr[1]/td/ul/li[3]/a").hover()
                  connecta.locator("//*[@id='tabelaMenu']/tbody/tr[1]/td/ul/li[3]/ul/li[1]/a").hover()
                  connecta.locator("a[href='controlerelatoriodeslocamento']").click()
                  connecta.locator("//*[@id='formfiltro']/fieldset/span[1]/button").click()
                  time.sleep(1)
                  filtro = pyautogui.locateOnScreen('filtro.png', confidence=0.7)
                  time.sleep(1)
                  pyautogui.moveTo(filtro)
                  time.sleep(1)
                  pyautogui.click()
                  pyautogui.write(placaReplaced)
                  time.sleep(1)
                  pyautogui.moveTo(x= filtro.left+50, y= filtro.top+45)
                  time.sleep(1)
                  pyautogui.click()
                  time.sleep(1)
                  dataIniti = connecta.locator("//*[@id='dtI']").click()
                  pyautogui.hotkey('ctrl','a')
                  pyautogui.press('del')
                  pyautogui.write(data_hour_for_connecta)
                  time.sleep(1)
                  pyautogui.press('Tab')
                  dataFinal = connecta.locator("//*[@id='dtF']").click()
                  pyautogui.hotkey('ctrl','a')
                  pyautogui.press('del')
                  pyautogui.write(data_hour_adicionada)
                  time.sleep(1)
                  pyautogui.press('Tab')
                  connecta.locator("//*[@id='formfiltro']/fieldset/div/button[2]").click()
                  try:
                    time.sleep(1)
                    identificado = pyautogui.locateOnScreen('sem_dados.png', confidence=0.9)
                    if(identificado == None):
                      newUrl = connecta.inner_html("//body/table/tbody/tr[5]/td/table/tbody/tr/td/table/tbody/tr[@class='even']")
                      newSoup = BeautifulSoup(newUrl, 'html.parser')
                      nome =  newSoup.find_all('td', attrs={'style': 'text-align: center'}, string=re.compile("[^a-z]"))
                      contador = 0
                      while contador < len(nome):
                       infrator = nome[0].text
                       if re.search("[a-zA-Z]", infrator):
                         nomeInfrator = infrator
                       contador+=1
                      page_multas.append([placaReplaced, renavan, results, nomeInfrator])
                      book.save(filename="C:/Users/wilk.silva/Downloads/Consulta dia "+formatData+".xlsx")   
                    else:
                      print('Condutor não identificado\n')
                      condutor = 'Condutor não identificado'
                      page_multas.append([placaReplaced, renavan, results, condutor])
                      book.save(filename="C:/Users/wilk.silva/Downloads/Consulta dia "+formatData+".xlsx")
                      print('Dados salvos sem o condutor')
                  except:
                    continue
              iContent+=1
        except:  
          try:
            time.sleep(1)
            localizador = pyautogui.locateOnScreen('debitos.png', confidence=0.9)
            if (localizador == None):
             select = newTab.locator('//*[@id="cmbTipoDebito"]').click()
             pyautogui.press('down')
             pyautogui.press('down')
             pyautogui.press('enter')
             time.sleep(1)
             infracoes = pyautogui.locateOnScreen('infracoes.png', confidence=0.9)
             if (infracoes == None):
              url = newTab.inner_html('//*[@id="div_servicos_Multas"]/table/tbody/tr[2]/td[2]')
              print("Existem dados")
              soup = BeautifulSoup(url, 'html.parser')
              debitos = soup.find_all('td', attrs={'class': False, 'width': False, 'colspan':False}, string=re.compile("às"))
              print('\n --- debitos identificados --- \n')
              print(debitos)
              iContent=0
              placaArray = [placa]
              placaParte1 = placa[0:3]
              placaParte2 =  placa[3:]
              placaReplaced = placaParte1+"-"+placaParte2
              book = openpyxl.load_workbook(filename="C:/Users/wilk.silva/Downloads/Consulta dia "+formatData+".xlsx")
              book.create_sheet("Multas")
              page_multas = book['Multas']
              while iContent < len(debitos):
                results = debitos[iContent].text
                if re.search("às", results):
                   print("\n\nLocal, data, e horário ------->>> "+results+"")
                   findData = re.search("[0-9][0-9]/[0-9][0-9]/[0-9][0-9][0-9][0-9]", results).group(0)
                   findHour = re.search("[0-9][0-9]:[0-9][0-9]", results).group(0)
                   data_hour_for_connecta = str(findData)+" "+str(findHour)
                   convertHour = str(findHour)
                   newHour = convertHour.replace(":", "")
                   aspsHour = newHour.replace(newHour,"'"+newHour+"'")
                   convertHour1 = aspsHour[1:3]
                   secondPartHour = aspsHour[3:5]
                   if (re.compile(convertHour1).match("00")):
                     acres = "0.1"
                     sumHour = int(convertHour1)+float(acres)
                   else:
                     sumHour = int(convertHour1)+1
                   horaAidicionada = str(sumHour)+":"+str(secondPartHour)
                   data_hour_adicionada = str(findData)+ " " +str(horaAidicionada)
                   connecta = context.new_page()
                   connecta.goto("http://www16.itrack.com.br/cmatrix/controlemonitoramento")
                   connecta.locator("input[name='usuario']").fill("GEAD")
                   connecta.locator("input[name='senha']").fill("33312976")
                   connecta.locator("button[name='Submit']").click()
                   connecta.locator("//*[@id='tabelaMenu']/tbody/tr[1]/td/ul/li[3]/a").hover()
                   connecta.locator("//*[@id='tabelaMenu']/tbody/tr[1]/td/ul/li[3]/ul/li[1]/a").hover()
                   connecta.locator("a[href='controlerelatoriodeslocamento']").click()
                   connecta.locator("//*[@id='formfiltro']/fieldset/span[1]/button").click()
                   time.sleep(1)
                   filtro = pyautogui.locateOnScreen('filtro.png', confidence=0.7)
                   time.sleep(1)
                   pyautogui.moveTo(filtro)
                   time.sleep(1)
                   pyautogui.click()
                   pyautogui.write(placaReplaced)
                   time.sleep(1)
                   pyautogui.moveTo(x= filtro.left+50, y= filtro.top+45)
                   time.sleep(1)
                   pyautogui.click()
                   dataIniti = connecta.locator("//*[@id='dtI']").click()
                   pyautogui.hotkey('ctrl','a')
                   pyautogui.press('del')
                   pyautogui.write(data_hour_for_connecta)
                   time.sleep(1)
                   pyautogui.press('Tab')
                   dataFinal = connecta.locator("//*[@id='dtF']").click()
                   pyautogui.hotkey('ctrl','a')
                   pyautogui.press('del')
                   pyautogui.write(data_hour_adicionada)
                   time.sleep(1)
                   pyautogui.press('Tab')
                   connecta.locator("//*[@id='formfiltro']/fieldset/div/button[2]").click()
                   try:
                     time.sleep(1)
                     identificado = pyautogui.locateOnScreen('sem_dados.png', confidence=0.9)
                     if(identificado == None):
                       newUrl = connecta.inner_html("//body/table/tbody/tr[5]/td/table/tbody/tr/td/table/tbody/tr[@class='even']")
                       newSoup = BeautifulSoup(newUrl, 'html.parser')
                       nome =  newSoup.find_all('td', attrs={'style': 'text-align: center'}, string=re.compile("[^a-z]"))
                       contador = 0
                       while contador < len(nome):
                        infrator = nome[0].text
                        if re.search("[a-zA-Z]", infrator):
                           nomeInfrator = infrator
                        contador+=1
                       page_multas.append([placaReplaced, renavan, results, nomeInfrator])
                       book.save(filename="C:/Users/wilk.silva/Downloads/Consulta dia "+formatData+".xlsx")  
                     else:
                       print('Condutor não identificado\n')
                       condutor = 'Condutor não identificado'
                       page_multas.append([placaReplaced, renavan, results, condutor])
                       book.save(filename="C:/Users/wilk.silva/Downloads/Consulta dia "+formatData+".xlsx")
                       print('Dados salvos sem o condutor')
                   except:
                     continue
                iContent+=1
          except: 
           print("Não existem dados") 
          continue
      except:
        print("Consulta concluída")
        tempoExec = time.time() -t1
        print("\nTempo de execução: {} segundos".format(tempoExec))
        continue  
      browser.close()