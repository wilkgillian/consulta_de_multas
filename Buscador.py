import pyautogui
import time
from datetime import datetime
from playwright.sync_api import sync_playwright
import pandas as pd
from bs4 import BeautifulSoup
import re
import openpyxl
from dotenv import load_dotenv
import os
load_dotenv()

t1 = time.time()
detran = "DETRAN"
dataAtual = datetime.today()
formatData = dataAtual.strftime("%d/%m/%Y").replace("/", "-")
planilha = pd.read_excel("base_de_dados/Controle - Frota.xlsx",
                         "Vencimento Documentação", skiprows=1, usecols=['PLACA', 'RENAVAN'])
try:
    planilha_formatada = pd.read_excel(
        "consultas/Consulta dia "+formatData+".xlsx")
except:
    planilha.to_excel("consultas/Consulta dia "+formatData+".xlsx")
    planilha_formatada = pd.read_excel(
        "consultas/Consulta dia "+formatData+".xlsx")
for index, row in planilha_formatada.iterrows():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, timeout=5000)
        context = browser.new_context()
        page = context.new_page()
        page.goto(os.environ['DET'])
        try:
            placa = row["PLACA"]
            renavan = row["RENAVAN"]
            try:
                page.locator('.closer').click()
            except:
                continue
            page.locator('#input_placa').fill(str(placa))
            page.locator('#input_renavam').fill(str(renavan).replace(".0", ""))
            with context.expect_event("page") as event_info:
                page.locator('//*[@id="formVeiculo"]/div[4]/input[2]').click()
            newTab = event_info.value
            book = openpyxl.load_workbook(
                filename="consultas/Consulta dia "+formatData+".xlsx")
            try:
                time.sleep(1)
                all_debits = pyautogui.locateOnScreen(
                    'all_debits.png', confidence=0.9)
                if (all_debits != None):
                    time.sleep(1)
                    select = newTab.locator('//*[@id="cmbTipoDebito"]').click()
                    pyautogui.press('down')
                    pyautogui.press('down')
                    pyautogui.press('enter')
                    time.sleep(1)
                    urlLicenciamento = newTab.inner_html(
                        '//*[@id="Integral"]/table/tbody')
                    soupLicenciamento = BeautifulSoup(
                        urlLicenciamento, 'html.parser')
                    licenciamentoArray = soupLicenciamento.find_all(
                        'td', attrs={'width': False, 'colspan': False})
                    liceContador = 0
                    print("--------------------------------------------------")
                    print(placa)
                    while liceContador < len(licenciamentoArray):
                        res = licenciamentoArray[liceContador].text
                        if re.search("Licenciamento", res):
                            licenciamentoRes = res
                            vencimentoLicenciamento = licenciamentoArray[liceContador+1].text
                            print("--->> "+placa+" "+licenciamentoRes +
                                  " vencimento: "+vencimentoLicenciamento+"")
                        liceContador += 1
                    try:
                        page_docs = book['Documentos']
                    except:
                        book.create_sheet("Documentos")
                        page_docs = book['Documentos']
                    page_docs.append(
                        [placa, renavan, licenciamentoRes, vencimentoLicenciamento])
                    book.save(
                        filename="consultas/Consulta dia "+formatData+".xlsx")
                else:
                    print("Sem licenciamento")
                localizador = pyautogui.locateOnScreen(
                    'debitos.png', confidence=0.9)
                if (localizador == None):
                    time.sleep(1)
                    select = newTab.locator('//*[@id="cmbTipoDebito"]').click()
                    pyautogui.press('down')
                    pyautogui.press('down')
                    pyautogui.press('enter')
                    time.sleep(1)
                    infracoes = pyautogui.locateOnScreen(
                        'infracoes.png', confidence=0.9)
                    if (infracoes == None):
                        time.sleep(1)
                        url = newTab.inner_html(
                            '//*[@id="div_servicos_Autuacoes"]/table/tbody/tr[2]/td[2]')
                        soup = BeautifulSoup(url, 'html.parser')
                        debitos = soup.find_all(
                            'td', attrs={'class': False, 'width': False, 'colspan': False})
                        iContent = 0
                        placaArray = [placa]
                        placaParte1 = placa[0:3]
                        placaParte2 = placa[3:]
                        placaReplaced = placaParte1+"-"+placaParte2
                        time.sleep(1)
                        try:
                            page_multas = book['Multas']
                        except:
                            book.create_sheet("Multas")
                            page_multas = book['Multas']
                        while iContent < len(debitos):
                            results = debitos[iContent].text
                            if re.search("às", results):
                                findData = re.search(
                                    "[0-9][0-9]/[0-9][0-9]/[0-9][0-9][0-9][0-9]", results).group(0)
                                findHour = re.search(
                                    "[0-9][0-9]:[0-9][0-9]", results).group(0)
                                data_hour_for_connecta = str(
                                    findData)+" "+str(findHour)
                                convertHour = str(findHour)
                                newHour = convertHour.replace(":", "")
                                aspsHour = newHour.replace(
                                    newHour, "'"+newHour+"'")
                                convertHour1 = aspsHour[1:3]
                                secondPartHour = aspsHour[3:5]
                                if (int(convertHour1) == 0):
                                    acres = '0.1'
                                    sumHour = int(convertHour1)+float(acres)
                                else:
                                    hora_acres = int(convertHour1)+1
                                    if hora_acres < 10:
                                        sumHour = '0'+str(hora_acres)
                                    else:
                                        sumHour = hora_acres
                                horaAidicionada = str(
                                    sumHour)+":"+str(secondPartHour)
                                data_hour_adicionada = str(
                                    findData) + " " + str(horaAidicionada)
                                time.sleep(1)
                                connecta = context.new_page()
                                connecta.goto(
                                    os.environ['CONNECTA'])
                                time.sleep(1)
                                connecta.locator(
                                    "input[name='usuario']").fill(os.environ['USER_NAME'])
                                connecta.locator(
                                    "input[name='senha']").fill(os.environ['PASSWORD'])
                                time.sleep(1)
                                connecta.locator(
                                    "button[name='Submit']").click()
                                connecta.locator(
                                    "//*[@id='tabelaMenu']/tbody/tr[1]/td/ul/li[3]/a").hover()
                                connecta.locator(
                                    "//*[@id='tabelaMenu']/tbody/tr[1]/td/ul/li[3]/ul/li[1]/a").hover()
                                connecta.locator(
                                    "a[href='controlerelatoriodeslocamento']").click()
                                connecta.locator(
                                    "//*[@id='formfiltro']/fieldset/span[1]/button").click()
                                time.sleep(1)
                                filtro = pyautogui.locateOnScreen(
                                    'filtro.png', confidence=0.7)
                                time.sleep(1)
                                pyautogui.moveTo(filtro)
                                time.sleep(1)
                                pyautogui.click()
                                pyautogui.write(placaReplaced)
                                time.sleep(1)
                                pyautogui.moveTo(
                                    x=filtro.left+50, y=filtro.top+45)
                                time.sleep(1)
                                pyautogui.click()
                                time.sleep(1)
                                dataIniti = connecta.locator(
                                    "//*[@id='dtI']").click()
                                time.sleep(1)
                                pyautogui.hotkey('ctrl', 'a')
                                pyautogui.press('del')
                                pyautogui.write(data_hour_for_connecta)
                                time.sleep(1)
                                pyautogui.press('Tab')
                                time.sleep(1)
                                dataFinal = connecta.locator(
                                    "//*[@id='dtF']").click()
                                pyautogui.hotkey('ctrl', 'a')
                                pyautogui.press('del')
                                pyautogui.write(data_hour_adicionada)
                                time.sleep(1)
                                pyautogui.press('Tab')
                                time.sleep(1)
                                connecta.locator(
                                    "//*[@id='formfiltro']/fieldset/div/button[2]").click()
                                try:
                                    time.sleep(1)
                                    identificado = pyautogui.locateOnScreen(
                                        'sem_dados.png', confidence=0.9)
                                    if(identificado == None):
                                        time.sleep(1)
                                        newUrl = connecta.inner_html(
                                            "//body/table/tbody/tr[5]/td/table/tbody/tr/td/table/tbody/tr[@class='even']")
                                        newSoup = BeautifulSoup(
                                            newUrl, 'html.parser')
                                        nome = newSoup.find_all(
                                            'td', attrs={'style': 'text-align: center'}, string=re.compile("[^a-z]"))
                                        contador = 0
                                        while contador < len(nome):
                                            infrator = nome[0].text
                                            if re.search("[a-zA-Z]", infrator):
                                                nomeInfrator = infrator
                                            contador += 1
                                        time.sleep(1)

                                        page_multas.append(
                                            [placaReplaced, renavan, results, nomeInfrator, detran])
                                        book.save(
                                            filename="consultas/Consulta dia "+formatData+".xlsx")
                                    else:
                                        print('Condutor não identificado\n')
                                        condutor = 'Condutor não identificado'
                                        time.sleep(1)
                                        page_multas.append(
                                            [placaReplaced, renavan, results, condutor, detran])
                                        book.save(
                                            filename="consultas/Consulta dia "+formatData+".xlsx")
                                        print('Dados salvos sem o condutor')
                                except:
                                    continue
                                connecta.locator(
                                    "//*[@id='li-sair']/a").click()
                                connecta.close()
                            iContent += 1
            except:
                try:
                    time.sleep(1)
                    localizador = pyautogui.locateOnScreen(
                        'debitos.png', confidence=0.9)
                    if (localizador == None):
                        time.sleep(1)
                        select = newTab.locator(
                            '//*[@id="cmbTipoDebito"]').click()
                        pyautogui.press('down')
                        pyautogui.press('down')
                        pyautogui.press('enter')
                        time.sleep(1)
                        infracoes = pyautogui.locateOnScreen(
                            'infracoes.png', confidence=0.9)
                        if (infracoes == None):
                            time.sleep(1)
                            url = newTab.inner_html(
                                '//*[@id="div_servicos_Multas"]/table/tbody/tr[2]/td[2]')
                            soup = BeautifulSoup(url, 'html.parser')
                            debitos = soup.find_all('td', attrs={
                                                    'class': False, 'width': False, 'colspan': False}, string=re.compile("às"))
                            iContent = 0
                            placaArray = [placa]
                            placaParte1 = placa[0:3]
                            placaParte2 = placa[3:]
                            placaReplaced = placaParte1+"-"+placaParte2
                            time.sleep(1)
                            try:
                                page_multas = book['Multas']
                            except:
                                book.create_sheet("Multas")
                                page_multas = book['Multas']
                            while iContent < len(debitos):
                                results = debitos[iContent].text
                                if re.search("às", results):
                                    print(
                                        "\n\nLocal, data, e horário ------->>> "+results+"")
                                    findData = re.search(
                                        "[0-9][0-9]/[0-9][0-9]/[0-9][0-9][0-9][0-9]", results).group(0)
                                    findHour = re.search(
                                        "[0-9][0-9]:[0-9][0-9]", results).group(0)
                                    data_hour_for_connecta = str(
                                        findData)+" "+str(findHour)
                                    convertHour = str(findHour)
                                    newHour = convertHour.replace(":", "")
                                    aspsHour = newHour.replace(
                                        newHour, "'"+newHour+"'")
                                    convertHour1 = aspsHour[1:3]
                                    secondPartHour = aspsHour[3:5]
                                    if (int(convertHour1) == 0):
                                        acres = '0.1'
                                        sumHour = int(
                                            convertHour1)+float(acres)
                                    else:
                                        hora_acres = int(convertHour1)+1
                                        if hora_acres < 10:
                                            sumHour = '0'+str(hora_acres)
                                        else:
                                            sumHour = hora_acres
                                    horaAidicionada = str(
                                        sumHour)+":"+str(secondPartHour)
                                    data_hour_adicionada = str(
                                        findData) + " " + str(horaAidicionada)
                                    time.sleep(1)
                                    connecta = context.new_page()
                                    time.sleep(1)
                                    connecta.goto(
                                        os.environ['CONNECTA'])
                                    connecta.locator(
                                        "input[name='usuario']").fill(os.environ['USER_NAME'])
                                    connecta.locator(
                                        "input[name='senha']").fill(os.environ['PASSWORD'])
                                    time.sleep(1)
                                    connecta.locator(
                                        "button[name='Submit']").click()
                                    connecta.locator(
                                        "//*[@id='tabelaMenu']/tbody/tr[1]/td/ul/li[3]/a").hover()
                                    connecta.locator(
                                        "//*[@id='tabelaMenu']/tbody/tr[1]/td/ul/li[3]/ul/li[1]/a").hover()
                                    connecta.locator(
                                        "a[href='controlerelatoriodeslocamento']").click()
                                    connecta.locator(
                                        "//*[@id='formfiltro']/fieldset/span[1]/button").click()
                                    time.sleep(1)
                                    filtro = pyautogui.locateOnScreen(
                                        'filtro.png', confidence=0.7)
                                    time.sleep(1)
                                    pyautogui.moveTo(filtro)
                                    time.sleep(1)
                                    pyautogui.click()
                                    pyautogui.write(placaReplaced)
                                    time.sleep(1)
                                    pyautogui.moveTo(
                                        x=filtro.left+50, y=filtro.top+45)
                                    time.sleep(1)
                                    pyautogui.click()
                                    time.sleep(1)
                                    dataIniti = connecta.locator(
                                        "//*[@id='dtI']").click()
                                    pyautogui.hotkey('ctrl', 'a')
                                    pyautogui.press('del')
                                    pyautogui.write(data_hour_for_connecta)
                                    time.sleep(1)
                                    pyautogui.press('Tab')
                                    time.sleep(1)
                                    dataFinal = connecta.locator(
                                        "//*[@id='dtF']").click()
                                    pyautogui.hotkey('ctrl', 'a')
                                    pyautogui.press('del')
                                    pyautogui.write(data_hour_adicionada)
                                    time.sleep(1)
                                    pyautogui.press('Tab')
                                    time.sleep(1)
                                    connecta.locator(
                                        "//*[@id='formfiltro']/fieldset/div/button[2]").click()
                                    try:
                                        time.sleep(1)
                                        identificado = pyautogui.locateOnScreen(
                                            'sem_dados.png', confidence=0.9)
                                        if(identificado == None):
                                            time.sleep(1)
                                            newUrl = connecta.inner_html(
                                                "//body/table/tbody/tr[5]/td/table/tbody/tr/td/table/tbody/tr[@class='even']")
                                            newSoup = BeautifulSoup(
                                                newUrl, 'html.parser')
                                            nome = newSoup.find_all(
                                                'td', attrs={'style': 'text-align: center'}, string=re.compile("[^a-z]"))
                                            contador = 0
                                            while contador < len(nome):
                                                infrator = nome[0].text
                                                if re.search("[a-zA-Z]", infrator):
                                                    nomeInfrator = infrator
                                                contador += 1
                                            time.sleep(1)
                                            page_multas.append(
                                                [placaReplaced, renavan, results, nomeInfrator, detran])
                                            book.save(
                                                filename="consultas/Consulta dia "+formatData+".xlsx")
                                        else:
                                            print('Condutor não identificado\n')
                                            condutor = 'Condutor não identificado'
                                            page_multas.append(
                                                [placaReplaced, renavan, results, condutor, detran])
                                            book.save(
                                                filename="consultas/Consulta dia "+formatData+".xlsx")
                                            print('Dados salvos sem o condutor')
                                    except:
                                        continue
                                    connecta.locator(
                                        "//*[@id='li-sair']/a").click()
                                    connecta.close()
                                iContent += 1
                except:
                    print("Não existem dados")
        except:
            print("Consulta concluída")
        browser.close()
tempoExec = time.time() - t1
print("\nTempo de execução: {} segundos".format(tempoExec))
